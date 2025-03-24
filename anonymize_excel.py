import re
import faker
import spacy
import pandas as pd
from openpyxl import load_workbook

fake = faker.Faker('it_IT')

name_mapping = {}
company_mapping = {}
email_mapping = {}
phone_mapping = {}
cf_mapping = {}
case_id_mapping = {}
account_id_mapping = {}
contract_id_mapping = {}
preventivo_id_mapping = {}
ritm_id_mapping = {}
task_id_mapping = {}
ods_id_mapping = {}
impianto_id_mapping = {}
conto_contrattuale_id_mapping = {}
bp_id_mapping = {}
mandato_id_mapping = {}
documento_id_mapping = {}
fm_id_mapping = {}
ordine_id_mapping = {}
cod_prat_utente_id_mapping = {}
asset_id_mapping = {}

# Carica il modello italiano di spaCy
nlp = spacy.load("it_core_news_lg")


def anonymize_name(name):
    if not name:
        return name
    if name not in name_mapping:
        name_mapping[name] = fake.name()
    return name_mapping[name]

def anonymize_company(company):
    if not company:
        return company
    if company not in company_mapping:
        company_mapping[company] = fake.company()
    return company_mapping[company]

def anonymize_email(email):
    if not email:
        return email
    if email not in email_mapping:
        username, domain = email.split('@') if '@' in email else (email, 'example.com')
        email_mapping[email] = f"{fake.user_name()}@{fake.domain_name()}"
    return email_mapping[email]

def anonymize_phone(phone):
    if not phone:
        return phone
    if phone not in phone_mapping:
        phone_mapping[phone] = fake.phone_number()
    return phone_mapping[phone]

def anonymize_cf(cf):
    if not cf:
        return cf
    if cf not in cf_mapping:
        cf_mapping[cf] = fake.ssn()
    return cf_mapping[cf]

def anonymize_id(id_val, mapping_dict, prefix='ANONYMIZED'):
    if not id_val:
        return id_val
    if id_val not in mapping_dict:
        mapping_dict[id_val] = f"{prefix}_{fake.random_number(digits=7)}"
    return mapping_dict[id_val]


def anonymize_text_spacy(text):
    if not text or not isinstance(text, str):
        return text

    doc = nlp(text)
    new_text_parts = []
    last_char_index = 0

    for ent in doc.ents:
        if ent.label_ == "PER":
            new_text_parts.append(text[last_char_index:ent.start_char])
            new_text_parts.append(anonymize_name(ent.text))
            last_char_index = ent.end_char
        else:
            pass

    new_text_parts.append(text[last_char_index:])
    text = "".join(new_text_parts)

    return text


def anonymize_email_phone_ids(text):
    if not text or not isinstance(text, str):
        return text

    # Anonymize emails
    for email in set(re.findall(r'\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\b', text)):
        text = text.replace(email, anonymize_email(email))

    # Anonymize phone numbers
    for phone in set(re.findall(r'\b(?:\+?39)?\s?\(?0?\)?\s?\d{3}\s?[-.\s]?\d{3}\s?[-.\s]?\d{4}\b', text)):
        text = text.replace(phone, anonymize_phone(phone))
    for phone in set(re.findall(r'\b0\d{9,10}\b', text)):
        text = text.replace(phone, anonymize_phone(phone))

    # Anonymize company names
    for company in set(re.findall(r'\b[A-Z][a-zA-Z0-9\s\.\&\,\-]+(?:S\.R\.L\.|S\.P\.A\.|S\.N\.C\.|SRL|SPA|SNC)\b', text)):
        text = text.replace(company, anonymize_company(company))
    for company in set(re.findall(r'\b[A-Z][a-zA-Z0-9\s\.\&\,\-]+(?:SRL|SPA|SNC)\b', text)):
        text = text.replace(company, anonymize_company(company))

    # Anonymize Case IDs
    for case_id in set(re.findall(r'\b0?9\d{6,7}\b', text)):
        text = text.replace(case_id, anonymize_id(case_id, case_id_mapping, 'CASE'))
    for case_id in set(re.findall(r'\bcase\s*0?9\d{6,7}\b', text, re.IGNORECASE)):
        text = text.replace(case_id, f"case {anonymize_id(case_id.split('case')[-1].strip(), case_id_mapping, 'CASE')}")
    for case_id in set(re.findall(r'\bcase\s*n\.\s*0?9\d{6,7}\b', text, re.IGNORECASE)):
        text = text.replace(case_id, f"case n. {anonymize_id(case_id.split('n.')[-1].strip(), case_id_mapping, 'CASE')}")
    for case_id in set(re.findall(r'\bCase\s*n\.\s*0?9\d{6,7}\b', text)):
        text = text.replace(case_id, f"Case n. {anonymize_id(case_id.split('n.')[-1].strip(), case_id_mapping, 'CASE')}")
    for case_id in set(re.findall(r'\bcase:\s*0?9\d{6,7}\b', text, re.IGNORECASE)):
        text = text.replace(case_id, f"case: {anonymize_id(case_id.split(':')[-1].strip(), case_id_mapping, 'CASE')}")

    # Anonymize Account IDs
    for account_id in set(re.findall(r'\b(?:20|25)0{3}\d{8}\b', text)):
        text = text.replace(account_id, anonymize_id(account_id, account_id_mapping, 'ACCOUNT'))
    for account_id in set(re.findall(r'\bac\s*(?:20|25)0{3}\d{8}\b', text, re.IGNORECASE)):
        text = text.replace(account_id, f"ac {anonymize_id(account_id.split('ac')[-1].strip(), account_id_mapping, 'ACCOUNT')}")
    for account_id in set(re.findall(r'\bca\s*(?:20|25)0{3}\d{8}\b', text, re.IGNORECASE)):
        text = text.replace(account_id, f"ca {anonymize_id(account_id.split('ca')[-1].strip(), account_id_mapping, 'ACCOUNT')}")
    for account_id in set(re.findall(r'\bconto contrattuale\s*(?:20|25)0{3}\d{8}\b', text, re.IGNORECASE)):
        text = text.replace(account_id, f"conto contrattuale {anonymize_id(account_id.split('contrattuale')[-1].strip(), account_id_mapping, 'ACCOUNT')}")

    # Anonymize Contract IDs
    for contract_id in set(re.findall(r'\b(?:3[56]0{2}|3)\d{8,10}\b', text)):
        text = text.replace(contract_id, anonymize_id(contract_id, contract_id_mapping, 'CONTRACT'))
    for contract_id in set(re.findall(r'\bcontratto\s*(?:3[56]0{2}|3)\d{8,10}\b', text, re.IGNORECASE)):
        text = text.replace(contract_id, f"contratto {anonymize_id(contract_id.split('contratto')[-1].strip(), contract_id_mapping, 'CONTRACT')}")

    # Anonymize Preventivo IDs
    for preventivo_id in set(re.findall(r'\b110\d{6}\b', text)):
        text = text.replace(preventivo_id, anonymize_id(preventivo_id, preventivo_id_mapping, 'PREVENTIVO'))
    for preventivo_id in set(re.findall(r'\bpreventivo\s*n\.?\s*110\d{6}\b', text, re.IGNORECASE)):
        text = text.replace(preventivo_id, f"preventivo n. {anonymize_id(preventivo_id.split('preventivo')[-1].strip(), preventivo_id_mapping, 'PREVENTIVO')}")

    # Anonymize RITM IDs
    for ritm_id in set(re.findall(r'\bRITM\d{7}\b', text)):
        text = text.replace(ritm_id, anonymize_id(ritm_id, ritm_id_mapping, 'RITM'))

    # Anonymize SCTASK IDs
    for task_id in set(re.findall(r'\bSCTASK\d{7}\b', text)):
        text = text.replace(task_id, anonymize_id(task_id, task_id_mapping, 'SCTASK'))
    for task_id in set(re.findall(r'\bTask SCTASK\d{7}\b', text)):
        text = text.replace(task_id, f"Task {anonymize_id(task_id.split('Task')[-1].strip(), task_id_mapping, 'SCTASK')}")

    # Anonymize ODS IDs
    for ods_id in set(re.findall(r'\b00095\d{8}\b', text)):
        text = text.replace(ods_id, anonymize_id(ods_id, ods_id_mapping, 'ODS'))
    for ods_id in set(re.findall(r'\bodl\s*95\d{8}\b', text, re.IGNORECASE)):
        text = text.replace(ods_id, f"odl {anonymize_id(ods_id.split('odl')[-1].strip(), ods_id_mapping, 'ODS')}")

    # Anonymize Impianto IDs
    for impianto_id in set(re.findall(r'\b400\d{7}\b', text)):
        text = text.replace(impianto_id, anonymize_id(impianto_id, impianto_id_mapping, 'IMPIANTO'))
    for impianto_id in set(re.findall(r'\bimpianto\s*400\d{7}\b', text, re.IGNORECASE)):
        text = text.replace(impianto_id, f"impianto {anonymize_id(impianto_id.split('impianto')[-1].strip(), impianto_id_mapping, 'IMPIANTO')}")

    # Anonymize Conto Contrattuale IDs
    for conto_contrattuale_id in set(re.findall(r'\b(?:2[05]0{3})\d{8}\b', text)):
        text = text.replace(conto_contrattuale_id, anonymize_id(conto_contrattuale_id, conto_contrattuale_id_mapping, 'CONTO_CONTRATTUALE'))

    # Anonymize BP IDs
    for bp_id in set(re.findall(r'\b(?:1[05]0{3})\d{6}\b', text)):
        text = text.replace(bp_id, anonymize_id(bp_id, bp_id_mapping, 'BP'))
    for bp_id in set(re.findall(r'\bBP-000\d{6}\b', text)):
        text = text.replace(bp_id, anonymize_id(bp_id, bp_id_mapping, 'BP'))

    # Anonymize Mandato IDs
    for mandato_id in set(re.findall(r'\b(?:003D\d{12}|A8S07\d{9})\d{4}\b', text)):
        text = text.replace(mandato_id, anonymize_id(mandato_id, mandato_id_mapping, 'MANDATO'))

    # Anonymize Documento IDs
    for documento_id in set(re.findall(r'\b20\d{14}\b', text)):
        text = text.replace(documento_id, anonymize_id(documento_id, documento_id_mapping, 'DOCUMENTO'))

    # Anonymize FM IDs
    for fm_id in set(re.findall(r'\bFM-\d{7}\b', text)):
        text = text.replace(fm_id, anonymize_id(fm_id, fm_id_mapping, 'FM'))

    # Anonymize Ordine IDs
    for ordine_id in set(re.findall(r'\b00095\d{8}\b', text)):
        text = text.replace(ordine_id, anonymize_id(ordine_id, ordine_id_mapping, 'ORDINE'))
    for ordine_id in set(re.findall(r'\bodl\s*95\d{8}\b', text, re.IGNORECASE)):
        text = text.replace(ordine_id, f"odl {anonymize_id(ordine_id.split('odl')[-1].strip(), ordine_id_mapping, 'ORDINE')}")

    # Anonymize COD_PRAT_UTENTE IDs
    for cod_prat_utente_id in set(re.findall(r'\bCA-0[6-9]\d{6}\b', text)):
        text = text.replace(cod_prat_utente_id, anonymize_id(cod_prat_utente_id, cod_prat_utente_id_mapping, 'COD_PRAT_UTENTE'))

    # Anonymize Asset IDs
    for asset_id in set(re.findall(r'\b02iWi\d{9}\b', text)):
        text = text.replace(asset_id, anonymize_id(asset_id, asset_id_mapping, 'ASSET'))

    return text


def anonymize_excel(input_path, output_path):
    # Carica il file Excel
    print(f"Caricamento del file Excel: {input_path}")
    
    # Utilizziamo pandas per leggere il file Excel
    try:
        # Leggi tutte le schede del file Excel
        excel_file = pd.ExcelFile(input_path)
        sheet_names = excel_file.sheet_names
        
        # Crea un writer per il file di output
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name in sheet_names:
                print(f"Elaborazione del foglio: {sheet_name}")
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                
                # Itera su ogni cella del DataFrame
                for row_idx, row in df.iterrows():
                    print(f"Elaborazione riga {row_idx + 1} di {len(df)}")
                    for col_name in df.columns:
                        cell_value = row[col_name]
                        
                        # Verifica se il valore è una stringa
                        if isinstance(cell_value, str):
                            # Applica le funzioni di anonimizzazione
                            anonymized_cell_value = anonymize_text_spacy(cell_value)
                            anonymized_cell_value = anonymize_email_phone_ids(anonymized_cell_value)
                            
                            # Aggiorna la cella nel DataFrame
                            df.at[row_idx, col_name] = anonymized_cell_value
                
                # Salva il foglio anonimizzato nel file di output
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"File Excel anonimizzato salvato in: {output_path}")
        
    except Exception as e:
        print(f"Errore durante l'elaborazione del file Excel: {str(e)}")


def anonymize_excel_preserve_formatting(input_path, output_path):
    """
    Versione alternativa che preserva la formattazione del file Excel originale.
    Utilizza openpyxl direttamente invece di pandas.
    """
    try:
        # Carica il workbook
        print(f"Caricamento del file Excel con preservazione della formattazione: {input_path}")
        wb = load_workbook(input_path)
        
        # Itera su ogni foglio
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"Elaborazione del foglio: {sheet_name}")
            
            # Itera su ogni cella
            total_rows = sheet.max_row
            for row_idx in range(1, total_rows + 1):
                print(f"Elaborazione riga {row_idx} di {total_rows}")
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    
                    # Verifica se il valore è una stringa
                    if isinstance(cell.value, str):
                        # Applica le funzioni di anonimizzazione
                        anonymized_value = anonymize_text_spacy(cell.value)
                        anonymized_value = anonymize_email_phone_ids(anonymized_value)
                        
                        # Aggiorna la cella
                        cell.value = anonymized_value
        
        # Salva il workbook anonimizzato
        wb.save(output_path)
        print(f"File Excel anonimizzato (con formattazione preservata) salvato in: {output_path}")
        
    except Exception as e:
        print(f"Errore durante l'elaborazione del file Excel: {str(e)}")


if __name__ == "__main__":
    input_excel_path = 'input.xlsx'  # Sostituisci con il percorso del tuo file Excel di input
    output_excel_path = 'output_anonymized.xlsx'
    
    # Scegli quale funzione utilizzare in base alle tue esigenze
    # anonymize_excel(input_excel_path, output_excel_path)  # Versione con pandas (più veloce ma potrebbe perdere formattazione)
    anonymize_excel_preserve_formatting(input_excel_path, output_excel_path)  # Versione che preserva la formattazione

